# -*- coding:utf-8 -*-
# Author:ChenZhipeng
# Time:2019/1/11 14:53

# https://openpyxl.readthedocs.io/en/stable/tutorial.html

import openpyxl
wb_total = openpyxl.load_workbook("次数统计.xlsx")             # 新建工作薄
wb_data = openpyxl.load_workbook('nginx_count_Jan.xlsx')         # 打开需要分析的工作薄
# sheet_creat.sheet_properties.tabColor = "1072BA"       x         # 给sheet表添加颜色
op_sheet = wb_data.sheetnames  # 获取分析的工作薄中的sheet名
print(op_sheet)
ws = wb_total.get_sheet_by_name("merge")
d1 = ws.cell(row=1, column=1, value="域名")   # 设置的第一列第一行的名称
for (i, j) in zip(range(1, 29), op_sheet):    # 根据所取数据的个数，仅需修改此处即可
    d = ws.cell(row=1, column=i, value=j)    # 根据sheet 命名列名
domains_list = []    # 用来存放每个sheet中每一行的数据，作为列表返回
page_list = [[], []]  # 第一个嵌套列表存放域名，第二个嵌套列表存放次数
for sheet in op_sheet:
    wb = wb_data.get_sheet_by_name(sheet)
    for row in wb.values:
        row = list(row)
        domains_list.append(row)
for data in domains_list:
    url = data[0]
    count = data[1]
    page_list[0].append(url)
    page_list[1].append(count)
merge_list = list(zip(*page_list))  # 行列置换
dict_merge = {}
for row in merge_list:
    dict_merge.setdefault(row[0], []).append(row[1])
data = []
data_list = []
for k, v in dict_merge.items():
    data.insert(0, k)
    for i in reversed(v):
        data.insert(1, i)
    data_list.append(data)
    data = []

for i in data_list:
    # print(i)
    ws.append(i)   # 按行添加数据到excel中

wb_total.save("次数统计.xlsx")


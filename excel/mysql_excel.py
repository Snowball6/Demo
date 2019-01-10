# -*- coding:utf-8 -*-
# Author:ChenZhipeng
# Time:2019/1/10 9:35

# # 定义日期格式
import datetime
import pymysql
import openpyxl
data_list = []
yes_time = datetime.datetime.today() + datetime.timedelta(-2)
yesterday_format = yes_time.strftime("%Y-%m-%d")
# 打开一个已存在的excel，如果新建使用workbook()
wb = openpyxl.load_workbook("nginx_count_Jan.xlsx")
# 以日期时间为sheet名，新建一个sheet
file = wb.create_sheet(yesterday_format)
# 获取excel工作薄中所有的sheet
sheet_list = wb.sheetnames
insert_sheet = sheet_list[-1]
# 操作上面新创建的sheet，默认操作第0个
ws = wb.get_sheet_by_name(insert_sheet)
# 命名单元格（即字段头）
ws["A1"] = "url"
ws["B1"] = "page_view"
ws["C1"] = "time"
conn = pymysql.connect(host="localhost", user="root", password="680034", db="test", port=3306)
cur = conn.cursor()
sql = "select * from nginx_information where time='%s' order by page_view desc" % yesterday_format
cur.execute(sql)
# 获取查到的所有数据
data = cur.fetchall()
for i in data:
    url = i[0]
    page_view = i[1]
    time = i[2]
    # 插入
    ws.append([url, page_view, time])
conn.close()
wb.save("nginx_count_Jan.xlsx")



